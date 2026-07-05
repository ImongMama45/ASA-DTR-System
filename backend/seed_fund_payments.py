import os
import re
from datetime import date, datetime

import django
from openpyxl import load_workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'dtr_project.settings')
django.setup()

from dtr_api.models import Employee, FundPayment

EXCEL_PATH = "../S.A FUND '25-'26.xlsx"
SHEET_NAMES = ["FUND 2025", "FUND 2026"]  # adjust if your workbook uses a single sheet
NAME_COL = 1          # column A holds employee names
HEADER_ROW = 1        # row containing the period headers (Jul-31, Aug-15, ...)
DATA_START_ROW = 2    # first row with actual employee data

RESIGNED_FILL_KEYWORDS = ("0000FF", "3B82F6", "0070C0", "0563C1", "4472C4", "0B5394")

def normalize(s):
    return re.sub(r'[^A-Z]', '', str(s).upper())

def match_employee(sheet_name, employees):
    sheet_name = str(sheet_name).strip()
    if not sheet_name or sheet_name.lower() == 'nan':
        return None
    if ',' in sheet_name:
        last, first = [x.strip() for x in sheet_name.split(',', 1)]
    else:
        parts = sheet_name.split(maxsplit=1)
        last = parts[0].strip()
        first = parts[1].strip() if len(parts) > 1 else ""

    for emp in employees:
        db_name = emp['name'].upper()
        if last.upper() in db_name:
            if first:
                first_part = first.split()[0].upper()
                if first_part in db_name:
                    return emp['id']
            else:
                return emp['id']
    return None

def cell_is_resigned(cell):
    fill = cell.fill
    if not fill or fill.fgColor is None:
        return False
    rgb = getattr(fill.fgColor, 'rgb', None)
    if not rgb or not isinstance(rgb, str):
        return False
    return any(key in rgb.upper() for key in RESIGNED_FILL_KEYWORDS)

def parse_period(header_cell):
    val = header_cell.value
    if val is None:
        return None
    if isinstance(val, date):
        d = val
    else:
        text = str(val).strip()
        d = None
        for fmt in ("%b-%d", "%b-%d-%Y", "%B-%d", "%b %d"):
            try:
                d = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                continue
        if d is None:
            return None
    cutoff = 1 if d.day <= 15 else 16
    return d.year, d.month, cutoff

def main():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    employees = list(Employee.objects.values('id', 'name'))

    FundPayment.objects.all().delete()

    count = 0
    skipped_unmatched = set()
    skipped_resigned = 0

    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found, skipping.")
            continue
        ws = wb[sheet_name]

        period_cols = {}
        for col_idx in range(NAME_COL + 1, ws.max_column + 1):
            header_cell = ws.cell(row=HEADER_ROW, column=col_idx)
            period = parse_period(header_cell)
            if period:
                period_cols[col_idx] = period

        if not period_cols:
            print(f"WARNING: no date headers detected on '{sheet_name}'. "
                  f"Check HEADER_ROW / NAME_COL constants.")

        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=NAME_COL)
            name = name_cell.value
            if not name or str(name).strip().lower() == 'nan':
                continue

            emp_id = match_employee(name, employees)
            if not emp_id:
                skipped_unmatched.add(str(name).strip())
                continue
            emp = Employee.objects.get(id=emp_id)

            for col_idx, (year, month, cutoff) in period_cols.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value

                if cell_is_resigned(cell):
                    skipped_resigned += 1
                    continue
                if val is None or str(val).strip().upper() == 'NEW' or str(val).strip() == '':
                    continue

                try:
                    amt = float(val)
                except (ValueError, TypeError):
                    continue

                if amt > 0:
                    FundPayment.objects.update_or_create(
                        employee=emp, year=year, month=month, cutoff=cutoff,
                        defaults={'amount': amt}
                    )
                    count += 1

    print(f"\nSeeded {count} payments.")
    if skipped_unmatched:
        print(f"\nSkipped {len(skipped_unmatched)} unmatched names (not in DB):")
        for n in sorted(skipped_unmatched):
            print(f"  - {n}")
    print(f"\nSkipped {skipped_resigned} resigned-period cells (blue-marked).")

if __name__ == "__main__":
    main()
