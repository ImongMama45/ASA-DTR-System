import os
import re
from datetime import date, datetime

import django
from openpyxl import load_workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'dtr_project.settings')
django.setup()

from dtr_api.models import Employee, FundPayment

EXCEL_PATH = "../S.A FUND '25-'26.xlsx"
SHEET_NAMES = ["FUND 2025", "FUND 2026"]  # adjust if your workbook uses a single sheet
NAME_COL = 1          # column A holds employee names
HEADER_ROW = 1        # row containing the period headers (Jul-31, Aug-15, ...)
DATA_START_ROW = 2    # first row with actual employee data

RESIGNED_FILL_KEYWORDS = ("0000FF", "3B82F6", "0070C0", "0563C1", "4472C4")

def inspect_fills(sheet_name, max_rows=15, max_cols=15):
    """Utility: print distinct fill colors found in a sheet region, so you
    can identify the correct 'resigned' blue code. Run this manually first
    if RESIGNED_FILL_KEYWORDS doesn't seem to be catching anything."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb[sheet_name]
    seen = set()
    for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row),
                             min_col=1, max_col=min(max_cols, ws.max_column)):
        for cell in row:
            rgb = getattr(cell.fill.fgColor, 'rgb', None) if cell.fill else None
            if rgb and isinstance(rgb, str) and rgb not in seen:
                seen.add(rgb)
                print(f"{cell.coordinate}: value={cell.value!r} fill={rgb}")

inspect_fills("FUND 2026", max_rows=50, max_cols=25)
