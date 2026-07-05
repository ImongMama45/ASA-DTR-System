import os
import django
import pandas as pd
from openpyxl import load_workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'dtr_project.settings')
django.setup()

from dtr_api.models import Employee, FundPayment

EXCEL_PATH = "../S.A FUND '25-'26.xlsx"
RESIGNED_FILL_KEYWORDS = ("0000FF", "3B82F6", "0070C0", "0563C1", "4472C4", "0B5394")

def match_employee(sheet_name, employees):
    sheet_name = str(sheet_name).strip()
    if not sheet_name or sheet_name.lower() == 'nan':
        return None
    if ',' in sheet_name:
        last, first = [x.strip() for x in sheet_name.split(',', 1)]
    else:
        parts = sheet_name.split(maxsplit=1)
        last = parts[0].strip()
        first = parts[1].strip() if len(parts) > 1 else ""
        
    for emp in employees:
        db_name = emp['name'].upper()
        if last.upper() in db_name:
            if first:
                first_part = first.split()[0].upper()
                if first_part in db_name:
                    return emp['id']
            else:
                return emp['id']
    return None

def cell_is_resigned(cell):
    fill = cell.fill
    if not fill or fill.fgColor is None:
        return False
    rgb = getattr(fill.fgColor, 'rgb', None)
    if not rgb or not isinstance(rgb, str):
        return False
    return any(key in rgb.upper() for key in RESIGNED_FILL_KEYWORDS)

def seed():
    FundPayment.objects.all().delete()
    employees = list(Employee.objects.values('id', 'name'))
    count = 0
    
    wb = load_workbook(EXCEL_PATH, data_only=True)
    
    # FUND 2026 Mapping (Jul 2025 - Jun 2026)
    # The columns start at C (index 3) because B is empty
    cols_2026 = [
        (2025, 7, 16), (2025, 8, 1), (2025, 8, 16), (2025, 9, 1), (2025, 9, 16), 
        (2025, 10, 1), (2025, 10, 16), (2025, 11, 1), (2025, 11, 16), (2025, 12, 1), (2025, 12, 16),
        (2026, 1, 1), (2026, 1, 16), (2026, 2, 1), (2026, 2, 16), (2026, 3, 1), (2026, 3, 16),
        (2026, 4, 1), (2026, 4, 16), (2026, 5, 1), (2026, 5, 16), (2026, 6, 1)
    ]
    
    ws = wb['FUND 2026']
    col_offset = 3 # openpyxl columns are 1-indexed, so C is 3
    
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        emp_id = match_employee(name, employees)
        if not emp_id: continue
        
        emp = Employee.objects.get(id=emp_id)
        
        for i, (year, month, cutoff) in enumerate(cols_2026):
            cell = ws.cell(row=row_idx, column=col_offset + i)
            val = cell.value
            
            if cell_is_resigned(cell):
                continue
                
            if val is None or str(val).strip().upper() == 'NEW' or str(val).strip() == '':
                continue
                
            try:
                amt = float(val)
                if amt > 0:
                    FundPayment.objects.update_or_create(
                        employee=emp, year=year, month=month, cutoff=cutoff,
                        defaults={'amount': amt}
                    )
                    count += 1
            except ValueError:
                pass
                
    print(f"Successfully seeded {count} payments.")

if __name__ == '__main__':
    seed()
