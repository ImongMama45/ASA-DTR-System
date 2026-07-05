import os
import django
from datetime import date, datetime
from openpyxl import load_workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'dtr_project.settings')
django.setup()

from dtr_api.models import Employee, FundPayment

EXCEL_PATH = "../S.A FUND '25-'26.xlsx"
SHEET_NAMES = ["FUND 2025", "FUND 2026"]
NAME_COL = 1
HEADER_ROW = 1
DATA_START_ROW = 2
RESIGNED_FILL_KEYWORDS = ("0000FF", "3B82F6", "0070C0", "0563C1", "4472C4", "0B5394")

def parse_period(header_cell):
    val = header_cell.value
    if val is None:
        return None
    if isinstance(val, date):
        d = val
    else:
        text = str(val).strip()
        d = None
        for fmt in ("%b-%d", "%b-%d-%Y", "%B-%d", "%b %d"):
            try:
                d = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                continue
        if d is None:
            return None
    cutoff = 1 if d.day <= 15 else 16
    return d.year, d.month, cutoff

def match_employee(sheet_name, employees):
    sheet_name = str(sheet_name).strip()
    if not sheet_name or sheet_name.lower() == 'nan':
        return None
    if ',' in sheet_name:
        last, first = [x.strip() for x in sheet_name.split(',', 1)]
    else:
        parts = sheet_name.split(maxsplit=1)
        last = parts[0].strip()
        first = parts[1].strip() if len(parts) > 1 else ""
    for emp in employees:
        db_name = emp['name'].upper()
        if last.upper() in db_name:
            if first:
                first_part = first.split()[0].upper()
                if first_part in db_name:
                    return emp['id']
            else:
                return emp['id']
    return None

def cell_is_resigned(cell):
    fill = cell.fill
    if not fill or fill.fgColor is None:
        return False
    rgb = getattr(fill.fgColor, 'rgb', None)
    if not rgb or not isinstance(rgb, str):
        return False
    return any(key in rgb.upper() for key in RESIGNED_FILL_KEYWORDS)

def get_name_and_data_start(ws, row_idx, name_col=1):
    col_a = ws.cell(row=row_idx, column=name_col).value
    col_b = ws.cell(row=row_idx, column=name_col + 1).value

    def looks_like_name_fragment(v):
        if v is None: return False
        if isinstance(v, (int, float, date, datetime)): return False
        text = str(v).strip()
        return text != "" and text.upper() != "NEW"

    if looks_like_name_fragment(col_b):
        full_name = f"{col_a}, {col_b}"
        data_start_col = name_col + 2
    else:
        full_name = col_a
        data_start_col = name_col + 1
    return full_name, data_start_col

def main():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    employees = list(Employee.objects.values('id', 'name'))
    
    FundPayment.objects.all().delete()
    
    count = 0
    skipped_unmatched = set()
    skipped_resigned = 0
    
    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        
        header_periods = []
        for col_idx in range(NAME_COL + 1, ws.max_column + 1):
            period = parse_period(ws.cell(row=HEADER_ROW, column=col_idx))
            if period:
                header_periods.append(period)

        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            name, data_start_col = get_name_and_data_start(ws, row_idx)
            if not name or str(name).strip().lower() == 'nan':
                continue
            emp_id = match_employee(name, employees)
            if not emp_id:
                skipped_unmatched.add(str(name).strip())
                continue
            emp = Employee.objects.get(id=emp_id)

            for i, (year, month, cutoff) in enumerate(header_periods):
                cell = ws.cell(row=row_idx, column=data_start_col + i)
                val = cell.value
                if cell_is_resigned(cell):
                    skipped_resigned += 1
                    continue
                if val is None or str(val).strip().upper() == 'NEW' or str(val).strip() == '':
                    continue
                try:
                    amt = float(val)
                except (ValueError, TypeError):
                    continue
                if amt > 0:
                    FundPayment.objects.update_or_create(
                        employee=emp, year=year, month=month, cutoff=cutoff,
                        defaults={'amount': amt}
                    )
                    count += 1
                    
    print(f"Seeded {count} payments using user's strict positional script.")
    
if __name__ == '__main__':
    main()
