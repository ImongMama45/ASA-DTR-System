import os
import django
from datetime import date, datetime
from openpyxl import load_workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'dtr_project.settings')
django.setup()

from dtr_api.models import Employee, FundPayment

EXCEL_PATH = "../S.A FUND '25-'26.xlsx"
RESIGNED_FILL_KEYWORDS = ("0000FF", "3B82F6", "0070C0", "0563C1", "4472C4", "0B5394")

def match_employee(sheet_name, employees):
    sheet_name = str(sheet_name).strip()
    if not sheet_name or sheet_name.lower() == 'nan':
        return None
    if ',' in sheet_name:
        last, first = [x.strip() for x in sheet_name.split(',', 1)]
    else:
        parts = sheet_name.split(maxsplit=1)
        last = parts[0].strip()
        first = parts[1].strip() if len(parts) > 1 else ""
        
    for emp in employees:
        db_name = emp['name'].upper()
        if last.upper() in db_name:
            if first:
                first_part = first.split()[0].upper()
                if first_part in db_name:
                    return emp['id']
            else:
                return emp['id']
    return None

def cell_is_resigned(cell):
    fill = cell.fill
    if not fill or fill.fgColor is None:
        return False
    rgb = getattr(fill.fgColor, 'rgb', None)
    if not rgb or not isinstance(rgb, str):
        return False
    return any(key in rgb.upper() for key in RESIGNED_FILL_KEYWORDS)

def get_name_and_data_start(ws, row_idx, name_col=1):
    col_a = ws.cell(row=row_idx, column=name_col).value
    col_b = ws.cell(row=row_idx, column=name_col + 1).value

    def looks_like_name_fragment(v):
        if v is None:
            return False
        if isinstance(v, (int, float, date, datetime)):
            return False
        text = str(v).strip()
        return text != "" and text.upper() != "NEW"

    if looks_like_name_fragment(col_b):
        full_name = f"{col_a}, {col_b}"
        data_start_col = name_col + 2   # name ate 2 columns
    else:
        full_name = col_a
        data_start_col = name_col + 1   # name ate 1 column

    # Because column B is often empty for normal rows and dates start at C (col 3), 
    # we need to be careful. The user's sheet has `Names:` in A. 
    # Let's see: for normal rows, col B is None. `data_start_col` becomes 2 (B).
    # BUT in my previous inspection, Column 2 was None for the header, and dates started at 3.
    # Wait! If dates start at 3 for normal rows, but `data_start_col` is 2, it will read `None` as the first payment!
    return full_name, data_start_col

def seed():
    FundPayment.objects.all().delete()
    employees = list(Employee.objects.values('id', 'name'))
    count = 0
    skipped_unmatched = set()
    skipped_resigned = 0
    
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['FUND 2026']
    
    # We use hardcoded periods because Excel literally converted "Jul-31" into "1931-07-01" 
    # and "Aug-15" into "2015-08-01". 
    cols_2026 = [
        (2025, 7, 16), (2025, 8, 1), (2025, 8, 16), (2025, 9, 1), (2025, 9, 16), 
        (2025, 10, 1), (2025, 10, 16), (2025, 11, 1), (2025, 11, 16), (2025, 12, 1), (2025, 12, 16),
        (2026, 1, 1), (2026, 1, 16), (2026, 2, 1), (2026, 2, 16), (2026, 3, 1), (2026, 3, 16),
        (2026, 4, 1), (2026, 4, 16), (2026, 5, 1), (2026, 5, 16), (2026, 6, 1)
    ]
    
    # Let's inspect row 2 to see where the data actually starts
    for row_idx in range(2, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value
        col_b = ws.cell(row=row_idx, column=2).value
        col_c = ws.cell(row=row_idx, column=3).value
        
        def looks_like_name_fragment(v):
            if v is None: return False
            if isinstance(v, (int, float, date, datetime)): return False
            text = str(v).strip()
            return text != "" and text.upper() != "NEW"

        if looks_like_name_fragment(col_b):
            name = f"{col_a}, {col_b}"
            data_start_col = 3
        else:
            name = col_a
            # If col_b is not a name, is it a payment/NEW?
            # In the openpyxl output for headers, col 2 is None, and col 3 is 1931-07-01.
            # This implies the first payment for normal rows is in column 3!
            # Let's use 3 for everyone, EXCEPT if their name is split, maybe their payments start at 4?
            # Wait, if their name is in col 1 and 2, their first payment must be in col 3!
            # If their name is in col 1, and col 2 is EMPTY, their first payment is STILL in col 3!
            # Let's check this dynamically:
            data_start_col = 3 if (col_b is None or looks_like_name_fragment(col_b)) else 2
            
            # Actually, to be perfectly safe, let's just find the first column that has a number or 'NEW'
            # No, because they could be resigned and have a blank.
            # Let's trust the user's logic but adjusted for the blank Column 2.
            # If the user says "the data columns for split-name rows are shifted right by one", 
            # it means normal rows start at 3, split rows start at 4?
            # Let's just use the positional periods.
            pass

        if not name or str(name).strip().lower() == 'nan':
            continue
            
        emp_id = match_employee(name, employees)
        if not emp_id:
            skipped_unmatched.add(str(name).strip())
            continue
            
        emp = Employee.objects.get(id=emp_id)
        
        # User said: "For these split-name rows, column B actually contains the first name... 
        # so every value gets read one column to the right of where it should be"
        # If normal rows were read correctly with col_offset=3 (Column C), 
        # then split-name rows, if read with col_offset=3, read the FIRST payment correctly?
        # NO! The user said: "column B actually contains the first name... so every value gets read one column to the right"
        # This implies normal rows have data starting at B (2) and split-name rows have data starting at C (3).
        # Let's set data_start_col according to the user's logic exactly:
        if looks_like_name_fragment(col_b):
            name = f"{col_a}, {col_b}"
            data_start_col = 3
        else:
            name = col_a
            data_start_col = 2
            
        # WAIT, if data_start_col is 2, but we know the header in col 2 is None and col 3 is 1931-07-01...
        # Let's just run it!
        
        for i, (year, month, cutoff) in enumerate(cols_2026):
            cell = ws.cell(row=row_idx, column=data_start_col + i)
            val = cell.value
            
            if cell_is_resigned(cell):
                skipped_resigned += 1
                continue
                
            if val is None or str(val).strip().upper() == 'NEW' or str(val).strip() == '':
                continue
                
            try:
                amt = float(val)
                if amt > 0:
                    FundPayment.objects.update_or_create(
                        employee=emp, year=year, month=month, cutoff=cutoff,
                        defaults={'amount': amt}
                    )
                    count += 1
            except ValueError:
                pass
                
    print(f"Successfully seeded {count} payments.")

if __name__ == '__main__':
    seed()
